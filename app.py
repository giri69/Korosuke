from flask import Flask, request, render_template
import PyPDF2
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'files[]' not in request.files:
        return 'No file part'
    
    files = request.files.getlist('files[]')
    texts = []

    for file in files:
        if file.filename == '':
            continue
        if file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ''
            for page_num in range(len(pdf_reader.pages)):
                text += pdf_reader.pages[page_num].extract_text()
            texts.append(text)

    # Print text from PDFs and write to Excel file
    excel_filename = 'extracted_text.xlsx'
    if os.path.exists(excel_filename):
        wb = load_workbook(excel_filename)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        start_row = 1

    for idx, text in enumerate(texts):
        print(f"Text from PDF {idx + 1}:")
        print(text)
        print("=" * 50)
        ws.cell(row=start_row+idx, column=1, value=f"Text from PDF {idx + 1}:")
        ws.cell(row=start_row+idx, column=2, value=text)

    wb.save(excel_filename)

    return f'Text extracted from PDFs has been saved in {excel_filename}'

if __name__ == '__main__':
    app.run(debug=True)
